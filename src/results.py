"""
This module contains functions for processing and saving classification results to an Excel file.
It provides functionality to save the results of image classification tasks performed by the application.
The results are saved in an Excel file, with each row representing a classification result.
The module also includes functionality to adjust the column widths in the Excel file for better visibility.
# THIS IS NEW.
"""

import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from image_processing import classify_images
from file_operations import create_directory, output_dir
import streamlit as st
from datetime import datetime
#import uuid
#import time

def save_results_to_excel(df_new, image_file_name=None):
    """
    Saves the classification results to an Excel file.

    This function takes a DataFrame containing the classification results and an optional image file name.
    It saves the results to an Excel file named 'Classification_Results.xlsx' in the output directory.
    If the Excel file already exists, it appends the new results to the existing file.
    It also adjusts the column widths in the Excel file for better visibility.

    Parameters:
    - df_new (pd.DataFrame): The DataFrame containing the new classification results.
    - image_file_name (str, optional): The name of the image file. Defaults to None.

    Raises:
    - PermissionError: If the Excel file is currently opened by the user.
    """
    try:
        # Ensure the output directory exists
        create_directory(output_dir)

        excel_file = os.path.join(output_dir, 'Classification_Results.xlsx')
        print(f"Image file name: {image_file_name}")
        print(f"df_new columns: {df_new.columns}")
        print(f"df_new shape: {df_new.shape}")

        df_new.loc[df_new.shape[0]-1, 'Filename'] = image_file_name # Add the image file name

        if os.path.exists(excel_file):
            df_old = pd.read_excel(excel_file)
            # Remove any rows in df_new that already exist in df_old
            df_new = df_new[~df_new.isin(df_old)].dropna()
            df = pd.concat([df_old, df_new], ignore_index=True)
        else:
            df = df_new

        # Insert an empty row after each row of data
        df = pd.concat([df, pd.DataFrame([pd.Series()], columns=df.columns)], ignore_index=True)

        # Save the DataFrame to the Excel file
        df.to_excel(excel_file, index=False)

    except Exception as e:
        print(f"Error saving results to Excel: {e}")

        # Auto update the column width to fit better for visibility
        book = openpyxl.load_workbook(excel_file)
        sheet = book.active

        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        book.save(excel_file)
    except Exception as e:
        print(f"An error occurred while saving results to Excel file: {e}")


def process_and_save_results(image, model_name, classification_data):
    """
    Processes and saves the classification results to an Excel file
    This function classifies the given image using the specified model, processes the results,
    and saves them to an Excel file. The filename for the Excel file is generated
    using a timestamp, class ID, and class name
    Parameters:
        - image (PIL.Image): The image to be classified.
        - model_name (str): The name of the model to use for classification.
        - classification_data (pd.DataFrame): The DataFrame containing the classification data
    Raises:
        - PermissionError: If the Excel file is currently opened by the user.
    """
    # Classify the image
    results = classify_images(image, model_name)

    # Save the classification results to an Excel file in the output directory
    if results.empty:
        print(f"No results to save for {image}")  # Debug print statement
    else:
        try:
            print(f"Saving results for {image} to Excel file")  # Debug print statement
            # Generate a timestamp for the current date and time
            datetime_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Assuming class_id and class_name are available from classification_data
            # For demonstration, let's assume we extract them from the first row of classification_data
            class_id = classification_data.iloc[0]['Class ID']
            class_name = classification_data.iloc[0]['Class Name']
            ext = "xlsx"  # Assuming the extension is xlsx for Excel file

            # Create the filename template
            filename_template = f"{datetime_stamp}_{class_id}_{class_name}.{ext}"

            # Use the filename template to create the image file name
            image_file_name = filename_template

            save_results_to_excel(results, image_file_name)
        except PermissionError:
            message = "Classification_Results.xlsx is currently opened by the user. Please, close it for the program to write out the classification results."
            print(message)
            st.error(message)
